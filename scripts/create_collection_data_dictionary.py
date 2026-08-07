"""Create a human-readable Excel data dictionary from the inferred schema."""
from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SOURCE = Path(r"D:\Digitalization\Project_Doc\realtime_collection_schema.json")
LATEST = SOURCE.parent / "realtime_latest_snapshot.json"
OUTPUT = SOURCE.parent / "Realtime_Collection_Column_Schema_Current.xlsx"


def autosize(sheet, widths=None):
    widths = widths or {}
    for index, column in enumerate(sheet.columns, 1):
        maximum = max((len(str(cell.value or "")) for cell in column), default=0)
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(index, min(max(maximum + 2, 10), 60))


def style_header(sheet):
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1769AA")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def main():
    payload = json.loads(SOURCE.read_text(encoding="utf-8"))
    latest_payload = json.loads(LATEST.read_text(encoding="utf-8"))
    latest_by_collection = {
        item["mapping"]["collection"]: item.get("document") or {}
        for item in latest_payload["documents"]
    }
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Collection Summary"
    summary.append([
        "Site Name", "Plant Name", "Customer Name", "Collection Name", "Latitude", "Longitude",
        "Collection Available", "Documents Sampled", "Latest Timestamp",
        "Column Count", "Columns Inside Collection",
    ])

    fields = workbook.create_sheet("Field Data Dictionary")
    fields.append([
        "Site Name", "Plant Name", "Customer Name", "Collection Name", "Latitude", "Longitude",
        "Column / Field Path", "Data Type", "Nullable", "Present In Samples",
        "Documents Sampled", "Example Values",
    ])

    current_tags = workbook.create_sheet("Exact Current Tags")
    current_tags.append([
        "Site Name", "Plant Name", "Customer Name", "Collection Name", "Latitude", "Longitude",
        "Exact Tag In Latest Document", "Current Value", "Inferred Data Type",
    ])

    matrix = workbook.create_sheet("Column Presence Matrix")
    all_fields = sorted({field for item in payload["collections"] for field in item["schema"]})
    matrix.append(["Site Name", "Plant Name", "Collection Name", *all_fields])

    for item in payload["collections"]:
        mapping = item["mapping"]
        schema = item["schema"]
        column_names = list(schema)
        latest = item.get("latestTimestamp")
        if isinstance(latest, dict):
            latest = latest.get("$date")
        summary.append([
            mapping["siteName"], mapping["plantName"], mapping["customerName"], mapping["collection"], mapping.get("lat"), mapping.get("lon"),
            "Yes" if item["exists"] else "No", item["sampledDocumentCount"], latest,
            len(column_names), ", ".join(column_names) if column_names else "No documents / schema unavailable",
        ])
        for field_name, definition in schema.items():
            examples = json.dumps(definition.get("examples", []), ensure_ascii=False)
            fields.append([
                mapping["siteName"], mapping["plantName"], mapping["customerName"], mapping["collection"], mapping.get("lat"), mapping.get("lon"),
                field_name, ", ".join(definition["bsonTypes"]),
                "Yes" if definition["nullable"] else "No", definition["presentInDocuments"],
                definition["sampleDocuments"], examples,
            ])
        latest_document = latest_by_collection.get(mapping["collection"], {})
        for tag, value in latest_document.items():
            definition = schema.get(tag, {})
            current_tags.append([
                mapping["siteName"], mapping["plantName"], mapping["customerName"], mapping["collection"],
                mapping.get("lat"), mapping.get("lon"), tag,
                json.dumps(value, ensure_ascii=False), ", ".join(definition.get("bsonTypes", [])),
            ])
        matrix.append([
            mapping["siteName"], mapping["plantName"], mapping["collection"],
            *[(", ".join(schema[field]["bsonTypes"]) if field in schema else "—") for field in all_fields],
        ])

    for sheet in (summary, fields, current_tags, matrix):
        style_header(sheet)
        sheet.sheet_view.showGridLines = False
    autosize(summary, {11: 100})
    autosize(fields, {12: 70})
    autosize(current_tags, {8: 45})
    autosize(matrix)
    for row in summary.iter_rows(min_row=2):
        row[10].alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(OUTPUT)
    print(OUTPUT)
    print(f"collections={len(payload['collections'])} uniqueFields={len(all_fields)} fieldRows={fields.max_row - 1}")


if __name__ == "__main__":
    main()
