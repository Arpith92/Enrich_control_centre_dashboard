"""Export mapped SCADA collection schemas and recent data without credentials."""
from __future__ import annotations

from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
import csv
import gzip
import json
from pathlib import Path
from typing import Any

from bson import json_util
from openpyxl import load_workbook
from pymongo import MongoClient

from sldc.config import settings

WORKBOOK = Path(r"D:\Digitalization\Project_Doc\Plant_wise_collection_details_with_latlong.xlsx")
OUTPUT_DIR = WORKBOOK.parent
SAMPLE_SIZE = 100


def bson_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "double"
    if isinstance(value, datetime):
        return "date"
    if isinstance(value, dict):
        return "object"
    if isinstance(value, list):
        return "array"
    return type(value).__name__


def flatten(document: dict, prefix: str = ""):
    for key, value in document.items():
        path = f"{prefix}.{key}" if prefix else str(key)
        yield path, value
        if isinstance(value, dict):
            yield from flatten(value, path)


def json_value(value: Any):
    return json.loads(json_util.dumps(value, json_options=json_util.RELAXED_JSON_OPTIONS))


def mappings() -> list[dict]:
    sheet = load_workbook(WORKBOOK, data_only=True, read_only=True).active
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({
            "serialNumber": row[0], "customerName": str(row[1] or "").strip(),
            "plantName": str(row[2] or "").strip(), "state": str(row[3] or "").strip(),
            "siteName": str(row[4] or "").strip(), "capacityAcMw": float(row[5] or 0),
            "capacityDcMwp": float(row[6] or 0),
            "commissioningDate": row[7].date().isoformat() if isinstance(row[7], datetime) else str(row[7] or ""),
            "collection": str(row[8] or "").strip(),
            "lat": float(row[9]) if row[9] is not None else None,
            "lon": float(row[10]) if row[10] is not None else None,
        })
    return rows


def inspect_collection(database, mapping: dict) -> dict:
    collection = database[mapping["collection"]]
    documents = list(collection.find().sort("_id", -1).limit(SAMPLE_SIZE))
    fields = defaultdict(lambda: {"types": Counter(), "present": 0, "examples": []})
    for document in documents:
        for path, value in flatten(document):
            entry = fields[path]
            entry["types"][bson_type(value)] += 1
            entry["present"] += 1
            example = json_value(value)
            if len(entry["examples"]) < 3 and example not in entry["examples"]:
                entry["examples"].append(example)
    schema = {}
    for path, entry in sorted(fields.items()):
        schema[path] = {
            "bsonTypes": sorted(entry["types"]),
            "nullable": entry["types"]["null"] > 0,
            "presentInDocuments": entry["present"],
            "sampleDocuments": len(documents),
            "examples": entry["examples"],
        }
    latest = documents[0] if documents else None
    return {
        "mapping": mapping,
        "exists": bool(documents),
        "sampledDocumentCount": len(documents),
        "latestTimestamp": json_value(latest.get("timestamp")) if latest and latest.get("timestamp") else None,
        "schema": schema,
        "latest": latest,
        "documents": documents,
    }


def main():
    if not settings.scada_mongodb_uri:
        raise SystemExit("SCADA_MONGODB_URI is not configured")
    rows = mappings()
    client = MongoClient(settings.scada_mongodb_uri, serverSelectionTimeoutMS=10_000, connectTimeoutMS=10_000)
    database_name = settings.scada_mongodb_database or client.get_default_database().name
    database = client[database_name]
    with ThreadPoolExecutor(max_workers=8) as executor:
        results = list(executor.map(lambda row: inspect_collection(database, row), rows))

    created_at = datetime.now(timezone.utc).isoformat()
    schema_output = {
        "schemaVersion": 1, "createdAtUtc": created_at, "sourceDatabase": database_name,
        "workbook": WORKBOOK.name, "sampleDocumentsPerCollection": SAMPLE_SIZE,
        "collectionCount": len(results),
        "collections": [{key: value for key, value in result.items() if key not in {"latest", "documents"}} for result in results],
    }
    (OUTPUT_DIR / "realtime_collection_schema.json").write_text(
        json.dumps(schema_output, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    latest_output = {
        "createdAtUtc": created_at, "sourceDatabase": database_name,
        "documents": [{"mapping": result["mapping"], "document": json_value(result["latest"])} for result in results],
    }
    (OUTPUT_DIR / "realtime_latest_snapshot.json").write_text(
        json.dumps(latest_output, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    dump_path = OUTPUT_DIR / "realtime_recent_data_dump.ndjson.gz"
    with gzip.open(dump_path, "wt", encoding="utf-8", newline="\n") as dump:
        for result in results:
            mapping = result["mapping"]
            for document in reversed(result["documents"]):
                dump.write(json.dumps({
                    "siteName": mapping["siteName"], "plantName": mapping["plantName"],
                    "collection": mapping["collection"], "document": json_value(document),
                }, ensure_ascii=False) + "\n")

    with (OUTPUT_DIR / "realtime_collection_inventory.csv").open("w", newline="", encoding="utf-8-sig") as output:
        columns = ["serialNumber", "siteName", "plantName", "customerName", "collection", "capacityAcMw", "capacityDcMwp", "exists", "sampledDocumentCount", "latestTimestamp"]
        writer = csv.DictWriter(output, fieldnames=columns)
        writer.writeheader()
        for result in results:
            inventory_row = {key: result["mapping"].get(key) for key in columns}
            inventory_row.update({
                "exists": result["exists"],
                "sampledDocumentCount": result["sampledDocumentCount"],
                "latestTimestamp": (result["latestTimestamp"] or {}).get("$date")
                if isinstance(result["latestTimestamp"], dict) else result["latestTimestamp"],
            })
            writer.writerow(inventory_row)

    missing = [result["mapping"]["collection"] for result in results if not result["exists"]]
    readme = f"""# Real-time SCADA schema and data dump

Generated: {created_at}
Source workbook: {WORKBOOK.name}
Source database: {database_name}
Mapped collections: {len(results)}
Collections with documents: {len(results) - len(missing)}
Missing/empty collections: {len(missing)}

## Files

- `realtime_collection_inventory.csv`: plant mapping, collection status, sampled count, and latest timestamp.
- `realtime_collection_schema.json`: inferred field paths and BSON-compatible types from up to {SAMPLE_SIZE} newest documents per collection.
- `realtime_latest_snapshot.json`: newest document from every mapped collection plus its plant metadata.
- `realtime_recent_data_dump.ndjson.gz`: up to {SAMPLE_SIZE} recent documents per collection in chronological order, one JSON object per line.

MongoDB ObjectIds and dates use MongoDB Extended JSON. Connection credentials are intentionally excluded.
Missing/empty collections: {', '.join(missing) if missing else 'None'}
"""
    (OUTPUT_DIR / "REALTIME_DATA_README.md").write_text(readme, encoding="utf-8")
    print(json.dumps({
        "outputDirectory": str(OUTPUT_DIR), "collections": len(results),
        "available": len(results) - len(missing), "missing": missing,
        "dumpDocuments": sum(len(result["documents"]) for result in results),
    }, indent=2))


if __name__ == "__main__":
    main()
