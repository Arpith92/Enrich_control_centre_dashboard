"""Read-only aggregation of the latest per-inverter SCADA values from MongoDB."""
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
import json
import logging
import os
import re
import threading
from functools import lru_cache
from zoneinfo import ZoneInfo

from pymongo.errors import ConfigurationError

from .config import settings

log = logging.getLogger(__name__)
TAG_RE = re.compile(
    r"^(?:Block[_ -]?0?(\d{1,3})[_ -])?INV(?:ERTER)?[_ -]?0?(\d{1,3})[_ -](Active[_ -]?Power|(?:Cumulative|Total)[_ -]?Generation)$", re.I,
)
DETAIL_TAG_RE = re.compile(
    r"^(?:Block[_ -]?0?(\d{1,3})[_ -])?INV(?:ERTER)?[_ -]?0?(\d{1,3})[_ -](Active[_ -]?Power|Daily[_ -]?Generation|(?:Cumulative|Total)[_ -]?Generation)$", re.I,
)
TIME_KEYS = ("timestamp_IST", "timestamp", "Timestamp", "time", "Time", "datetime", "DateTime", "createdAt", "updatedAt")
PLANT_TOTAL_TAGS = {
    "active_power": {"activepower", "plantactivepower", "totalactivepower", "acactivepower", "acpower", "currentpower"},
    "daily_generation": {"dailygeneration", "todaygeneration", "daygeneration", "generationtoday"},
    "cumulative_generation": {"cumulativegeneration", "totalgeneration", "lifetimegeneration", "energytotal"},
}
BHOKAR_LIVE_COLLECTIONS = [
    "B1_Jugai_LIVE", "B2_Jagdeesh_LIVE", "B3_Supriya_LIVE",
    "B4_Padmavati_LIVE", "B5_SoundCasting_LIVE", "B6_IMP_LIVE",
    "B7_Suyash_LIVE", "B8_Veersha_LIVE", "B9_Omya_LIVE",
]
UMRI_LIVE_COLLECTIONS = [
    "U1_WHF_LIVE", "U2_WIF_LIVE", "U3_Klassic_LIVE", "U4_Marvelous_LIVE",
    "U5_Haldiram_LIVE", "U6_Parakh_LIVE", "U7_PV_Sons_LIVE", "U9_WHF_2_LIVE",
]
DEFAULT_LIVE_COLLECTIONS = {
    "Bhokar": BHOKAR_LIVE_COLLECTIONS,
    "Umri": UMRI_LIVE_COLLECTIONS,
    "PGCIL": ["PGCIL_LIVE"],
}
WORKBOOK_SITE_ALIASES = {"Bhokar - I": "Bhokar", "Polangal": "NLC Poolangal"}


def _coordinate(value, limit):
    if value is None:
        return None
    coordinate = float(value)
    while abs(coordinate) > limit:
        coordinate /= 10
    return coordinate


@lru_cache(maxsize=1)
def workbook_collection_mapping() -> tuple[dict, dict]:
    path = settings.scada_mapping_workbook
    if not path.exists():
        return {}, {}
    from openpyxl import load_workbook
    collections: dict[str, list[str]] = {}
    metadata = {}
    sheet = load_workbook(path, data_only=True, read_only=True).active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        collection = str(row[8] or "").strip() if len(row) > 8 else ""
        if not collection:
            continue
        workbook_site = str(row[4] or "").strip()
        site = WORKBOOK_SITE_ALIASES.get(workbook_site, workbook_site)
        collections.setdefault(site, []).append(collection)
        metadata[collection.casefold()] = {
            "customerName": str(row[1] or "").strip(), "plantName": str(row[2] or "").strip(),
            "state": str(row[3] or "").strip(), "siteName": site,
            "ac": float(row[5] or 0), "dc": float(row[6] or 0),
            "lat": _coordinate(row[9], 90) if len(row) > 9 else None,
            "lon": _coordinate(row[10], 180) if len(row) > 10 else None,
        }
    return collections, metadata


def _json_mapping(raw: str) -> dict:
    try:
        value = json.loads(raw or "{}")
        return value if isinstance(value, dict) else {}
    except (TypeError, json.JSONDecodeError):
        log.warning("Invalid SCADA JSON mapping; ignoring it")
        return {}


def site_configuration() -> dict[str, dict]:
    workbook_collections, _ = workbook_collection_mapping()
    collections = {name: list(values) for name, values in workbook_collections.items()}
    collections.update(_json_mapping(settings.scada_site_collections))
    databases = _json_mapping(settings.scada_site_databases)
    names = set(collections) | set(databases)
    # Simple per-site variables are convenient in .env and override JSON mappings.
    for key, value in os.environ.items():
        match = re.fullmatch(r"SCADA_(.+)_COLLECTIONS", key)
        if match and key != "SCADA_SITE_COLLECTIONS":
            names.add(match.group(1).replace("_", " ").title())
    result = {}
    if settings.scada_mongodb_uri:
        for site, defaults in DEFAULT_LIVE_COLLECTIONS.items():
            names.add(site)
            collections.setdefault(site, defaults)
    for name in names:
        env_key = re.sub(r"[^A-Z0-9]+", "_", name.upper()).strip("_")
        raw = os.getenv(f"SCADA_{env_key}_COLLECTIONS")
        configured = raw.split(",") if raw is not None else collections.get(name, [])
        if isinstance(configured, str):
            configured = configured.split(",")
        collection_names = [str(item).strip() for item in configured if str(item).strip()]
        if name.casefold() in {"bhokar", "umri"}:
            live_collections = [item for item in collection_names if item.upper().endswith("_LIVE")]
            if live_collections:
                collection_names = live_collections
            else:
                collection_names = [
                    re.sub(r"_(?:daily|monthly|yearly)$", "", item, flags=re.I) + "_LIVE"
                    for item in collection_names
                ]
            collection_names = sorted(set(collection_names), key=lambda item: int(re.match(r"[BU](\d+)", item, re.I).group(1)))
        database = os.getenv(f"SCADA_{env_key}_DATABASE", databases.get(name, settings.scada_mongodb_database))
        if collection_names:
            result[name] = {"database": database, "collections": collection_names}
    return result


def _number(value):
    if isinstance(value, dict):
        for key in ("value", "Value", "val", "reading"):
            if key in value:
                return _number(value[key])
        return None
    if isinstance(value, bool):
        return None
    try:
        number = float(str(value).replace(",", "").strip())
        return number if number == number else None
    except (TypeError, ValueError):
        return None


def _positive_generation_value(value):
    """Normalize reversed meter signs while dropping common Modbus fault sentinels."""
    number = _number(value)
    if number is None:
        return None
    if number <= -2_147_000 or -2148 <= number <= -2146:
        return None
    return abs(number)


def _lifetime_to_mwh(value, *, small_values_are_gwh=False):
    """Normalize mixed lifetime counters: large counters are kWh, smaller counters are already MWh."""
    number = _positive_generation_value(value)
    if number is None:
        return None
    if small_values_are_gwh and number < 100_000:
        return number * 1000
    return number * settings.scada_generation_to_mwh if number >= 100_000 else number


def extract_inverter_values(document: dict) -> dict[str, dict[int, float]]:
    values = {"active_power": {}, "cumulative_generation": {}}

    def visit(node):
        if isinstance(node, dict):
            # Also accept common {tag/name: ..., value: ...} SCADA record shapes.
            tag = next((node.get(key) for key in ("tag", "Tag", "tagName", "TagName", "name", "Name") if node.get(key)), None)
            if tag:
                consume(str(tag), next((node.get(key) for key in ("value", "Value", "val", "reading") if key in node), None))
            for key, value in node.items():
                consume(str(key), value)
                if isinstance(value, (dict, list)):
                    visit(value)
        elif isinstance(node, list):
            for item in node:
                visit(item)

    def consume(tag, raw):
        match = TAG_RE.fullmatch(tag.strip())
        if not match:
            return
        block = int(match.group(1) or 0)
        inverter = int(match.group(2))
        if inverter < 1:
            return
        inverter_key = block * 1000 + inverter
        number = _positive_generation_value(raw)
        if number is None:
            return
        kind = "active_power" if re.sub(r"[^a-z]", "", match.group(3).lower()) == "activepower" else "cumulative_generation"
        values[kind][inverter_key] = number

    visit(document)
    return values


def extract_plant_totals(document: dict) -> dict[str, float | None]:
    """Read aggregate plant fields used by compact *_LIVE documents."""
    totals = {key: None for key in PLANT_TOTAL_TAGS}

    def visit(node):
        if isinstance(node, dict):
            tag = next((node.get(key) for key in ("tag", "Tag", "tagName", "TagName", "name", "Name") if node.get(key)), None)
            if tag:
                consume(str(tag), next((node.get(key) for key in ("value", "Value", "val", "reading") if key in node), None))
            for key, value in node.items():
                consume(str(key), value)
                if isinstance(value, (dict, list)):
                    visit(value)
        elif isinstance(node, list):
            for item in node:
                visit(item)

    def consume(tag, raw):
        normalized = re.sub(r"[^a-z0-9]", "", tag.casefold())
        for kind, names in PLANT_TOTAL_TAGS.items():
            if totals[kind] is None and normalized in names:
                totals[kind] = _number(raw)

    visit(document)
    return totals


def _document_time(document: dict):
    for key in TIME_KEYS:
        value = document.get(key)
        if isinstance(value, datetime):
            if value.tzinfo:
                return value
            zone = ZoneInfo("Asia/Kolkata") if key == "timestamp_IST" else timezone.utc
            return value.replace(tzinfo=zone)
        if isinstance(value, str):
            try:
                parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
                if parsed.tzinfo:
                    return parsed
                zone = ZoneInfo("Asia/Kolkata") if key == "timestamp_IST" else timezone.utc
                return parsed.replace(tzinfo=zone)
            except ValueError:
                pass
    object_id = document.get("_id")
    return getattr(object_id, "generation_time", None)


def _serializable_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


class ScadaRealtimeReader:
    def __init__(self):
        self.client = None
        self._details_locks = {}
        self._details_locks_guard = threading.Lock()

    def _site_lock(self, site: str):
        with self._details_locks_guard:
            return self._details_locks.setdefault(site, threading.Lock())

    def _client(self):
        if not settings.scada_mongodb_uri:
            return None
        if self.client is None:
            from pymongo import MongoClient
            self.client = MongoClient(settings.scada_mongodb_uri, serverSelectionTimeoutMS=5000, connectTimeoutMS=5000)
        return self.client

    def latest(self) -> dict:
        configured = site_configuration()
        if not settings.scada_mongodb_uri or not configured:
            return {"configured": False, "sites": [], "message": "SCADA MongoDB is not configured"}
        client = self._client()
        try:
            uri_database = client.get_default_database().name
        except (AttributeError, TypeError, ConfigurationError):
            uri_database = ""
        now = datetime.now(timezone.utc)
        sites = []
        for site, config in configured.items():
            power, generation, power_inverters, generation_inverters, inverter_total, timestamps, errors = 0.0, 0.0, set(), set(), 0, [], []
            database_name = config["database"] or uri_database
            if not database_name:
                sites.append({
                    "name": site, "live": False, "currentMw": None,
                    "cumulativeGenerationMWh": None, "inverterCount": 0,
                    "collectionCount": len(config["collections"]), "timestamp": None,
                    "errors": ["Database missing: include it in SCADA_MONGODB_URI"],
                })
                continue
            for collection_name in config["collections"]:
                try:
                    # Deliberately read-only: this client is never passed to the SLDC writer.
                    document = client[database_name][collection_name].find_one(sort=[("_id", -1)])
                    if not document:
                        continue
                    timestamp = _document_time(document)
                    if timestamp:
                        timestamps.append(timestamp)
                        if (now - timestamp.astimezone(timezone.utc)).total_seconds() > settings.scada_max_age_seconds:
                            continue
                    values = extract_inverter_values(document)
                    totals = extract_plant_totals(document)
                    document_power = sum(values["active_power"].values()) if values["active_power"] else totals["active_power"]
                    document_generation = sum(values["cumulative_generation"].values()) if values["cumulative_generation"] else totals["cumulative_generation"]
                    document_power = _positive_generation_value(document_power)
                    document_generation = _positive_generation_value(document_generation)
                    if document_power is not None:
                        power += document_power
                    if document_generation is not None:
                        generation += document_generation
                    power_inverters.update(values["active_power"])
                    generation_inverters.update(values["cumulative_generation"])
                    if document_power is not None and not values["active_power"]:
                        power_inverters.add(len(power_inverters) + 1)
                    if document_generation is not None and not values["cumulative_generation"]:
                        generation_inverters.add(len(generation_inverters) + 1)
                    inverter_total += len(set(values["active_power"]) | set(values["cumulative_generation"]))
                except Exception as exc:  # isolate one inaccessible collection from the rest of a site
                    errors.append(f"{collection_name}: {exc}")
            sites.append({
                "name": site,
                "live": bool(power_inverters),
                "communicationIssue": not bool(power_inverters),
                "sampleType": "1-minute average",
                "currentMw": round(power * settings.scada_power_to_mw, 3) if power_inverters else None,
                "cumulativeGenerationMWh": round(generation * settings.scada_generation_to_mwh, 3) if generation_inverters else None,
                "inverterCount": inverter_total,
                "collectionCount": len(config["collections"]),
                "timestamp": max(timestamps).isoformat() if timestamps else None,
                "errors": errors,
            })
        return {"configured": True, "sites": sites}

    def site_details(self, requested_site: str) -> dict | None:
        configured = site_configuration()
        site = next((name for name in configured if name.casefold() == requested_site.casefold()), None)
        if not site or not settings.scada_mongodb_uri:
            return None
        # Serialize reads for the same site, but always fetch the newest document.
        # The *_LIVE collections update every second and must never be API-cached.
        with self._site_lock(site):
            return self._load_site_details(site, configured[site])

    def _load_site_details(self, site: str, config: dict) -> dict:
        client = self._client()
        try:
            uri_database = client.get_default_database().name
        except (AttributeError, TypeError, ConfigurationError):
            uri_database = ""
        database_name = config["database"] or uri_database
        if not database_name:
            return {"name": site, "plants": [], "error": "SCADA database is not configured"}

        def read_plant(collection_name):
            try:
                document = client[database_name][collection_name].find_one(sort=[("_id", -1)])
                if not document:
                    return {"collection": collection_name, "name": _plant_name(collection_name), "available": False}
                timestamp = _document_time(document)
                inverters: dict[int, dict] = {}
                parameters = {}
                raw_tags = {}

                def consume_detail(tag, raw):
                    match = DETAIL_TAG_RE.fullmatch(str(tag).strip())
                    if not match or int(match.group(2)) < 1:
                        return False
                    block = int(match.group(1) or 0)
                    inverter_number = int(match.group(2))
                    inverter_key = block * 1000 + inverter_number
                    inverter_label = f"Block {block} Inv {inverter_number}" if block else inverter_number
                    inverter = inverters.setdefault(inverter_key, {"inverter": inverter_label})
                    value = _positive_generation_value(raw)
                    kind = re.sub(r"[^a-z]", "", match.group(3).lower())
                    if kind == "activepower":
                        inverter["activePowerRaw"] = value
                        inverter["activePowerMw"] = round(value * settings.scada_power_to_mw, 3) if value is not None else None
                    elif kind == "dailygeneration":
                        inverter["dailyGenerationMWh"] = round(value * settings.scada_generation_to_mwh, 3) if value is not None else None
                    else:
                        lifetime_mwh = _lifetime_to_mwh(value, small_values_are_gwh=collection_name.casefold() == "nlc_live")
                        inverter["cumulativeGenerationMWh"] = round(lifetime_mwh, 3) if lifetime_mwh is not None else None
                    return True

                def visit_detail(node, path=""):
                    if isinstance(node, dict):
                        tag = next((node.get(key) for key in ("tag", "Tag", "tagName", "TagName", "name", "Name") if node.get(key)), None)
                        if tag:
                            raw = next((node.get(key) for key in ("value", "Value", "val", "reading") if key in node), None)
                            consume_detail(tag, raw)
                        for key, raw in node.items():
                            if key == "_id" or key in TIME_KEYS:
                                continue
                            key_path = f"{path}.{key}" if path else str(key)
                            raw_tags[key_path] = _serializable_value(raw)
                            matched = consume_detail(key, raw)
                            if isinstance(raw, (dict, list)):
                                visit_detail(raw, key_path)
                            elif not matched:
                                parameters[key_path] = raw
                    elif isinstance(node, list):
                        for index, item in enumerate(node):
                            visit_detail(item, f"{path}[{index}]")

                visit_detail(document)
                rows = [inverters[number] for number in sorted(inverters)]
                totals = extract_plant_totals(document)
                active_values = [row["activePowerMw"] for row in rows if row.get("activePowerMw") is not None]
                daily_values = [row["dailyGenerationMWh"] for row in rows if row.get("dailyGenerationMWh") is not None]
                cumulative_values = [row["cumulativeGenerationMWh"] for row in rows if row.get("cumulativeGenerationMWh") is not None]
                current_mw = sum(active_values) if active_values else (
                    round(_positive_generation_value(totals["active_power"]) * settings.scada_power_to_mw, 3) if _positive_generation_value(totals["active_power"]) is not None else None
                )
                daily_mwh = sum(daily_values) if daily_values else (
                    round(_positive_generation_value(totals["daily_generation"]) * settings.scada_generation_to_mwh, 3) if _positive_generation_value(totals["daily_generation"]) is not None else None
                )
                cumulative_mwh = sum(cumulative_values) if cumulative_values else (
                    round(_lifetime_to_mwh(totals["cumulative_generation"]), 3) if _lifetime_to_mwh(totals["cumulative_generation"]) is not None else None
                )
                _, workbook_metadata = workbook_collection_mapping()
                metadata = workbook_metadata.get(collection_name.casefold(), {})
                return {
                    "collection": collection_name,
                    "name": metadata.get("plantName") or _plant_name(collection_name),
                    **metadata,
                    # A collection is communicating when its latest document can
                    # be read. Zero/missing active power is an operational value,
                    # not a collection communication failure.
                    "available": bool(document),
                    "dataAvailable": current_mw is not None,
                    "timestamp": timestamp.isoformat() if timestamp else None,
                    "parameters": parameters,
                    "rawTags": raw_tags,
                    "inverters": rows,
                    "currentMw": current_mw,
                    "dailyGenerationMWh": daily_mwh,
                    "cumulativeGenerationMWh": cumulative_mwh,
                }
            except Exception as exc:
                log.warning("%s SCADA collection %s unavailable: %s", site, collection_name, exc)
                return {
                    "collection": collection_name, "name": _plant_name(collection_name),
                    "available": False, "error": "Cloud SCADA temporarily unavailable; retrying on the next one-minute refresh",
                }

        with ThreadPoolExecutor(max_workers=min(4, len(config["collections"]))) as executor:
            plants = list(executor.map(read_plant, config["collections"]))

        for plant in plants:
            if plant.get("available"):
                plant["stale"] = False

        cumulative_totals = [plant["cumulativeGenerationMWh"] for plant in plants if plant.get("cumulativeGenerationMWh") is not None]
        result = {
            "name": site,
            "timestamp": max((plant.get("timestamp") for plant in plants if plant.get("timestamp")), default=None),
            "currentMw": round(sum(plant.get("currentMw") or 0 for plant in plants), 3),
            "dailyGenerationMWh": round(sum(plant.get("dailyGenerationMWh") or 0 for plant in plants), 3),
            "cumulativeGenerationMWh": round(sum(cumulative_totals), 3) if cumulative_totals else None,
            "plants": plants,
        }
        return result

    def plant_details(self, requested_site: str, requested_collection: str) -> dict | None:
        configured = site_configuration()
        site = next((name for name in configured if name.casefold() == requested_site.casefold()), None)
        if not site:
            return None
        collection = next((
            name for name in configured[site]["collections"]
            if name.casefold() == requested_collection.casefold()
        ), None)
        if not collection:
            return None
        details = self.site_details(site)
        if not details:
            return None
        return next((plant for plant in details.get("plants", []) if plant["collection"] == collection), None)


def _plant_name(collection_name: str) -> str:
    name = re.sub(r"^[BU]\d+[_ -]*", "", collection_name, flags=re.I)
    normalized = re.sub(r"[^a-z0-9]", "", name.casefold())
    aliases = {
        "jagdeeshlive": "Jagadeesh",
        "suyashlive": "Suyesh",
        "soundcastinglive": "Sound Castings",
        "veershalive": "Veeresha",
        "wiflive": "WIF",
        "whflive": "WHF-1",
        "whf2live": "WHF-2",
        "klassiclive": "Klassic Wheels",
        "marvelouslive": "Marvelous",
        "haldiramlive": "Haldiram",
        "parakhlive": "Parakh",
        "pvsonslive": "PV Sons",
    }
    if normalized in aliases:
        return aliases[normalized]
    return re.sub(r"[_-]+", " ", name).strip() or collection_name


scada_reader = ScadaRealtimeReader()
