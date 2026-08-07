"""Single-port server for the React control centre and integrated MH SLDC Scout."""
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta
import hashlib
import math
from io import BytesIO
from pathlib import Path
import re
import requests

from fastapi import FastAPI, HTTPException, Query, Response
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field

from sldc.database import db
from sldc.config import settings
from sldc.parser import TARGET_STATIONS
from sldc.scheduler import collector
from sldc.scada_realtime import scada_reader

ROOT = Path(__file__).resolve().parent
DIST = ROOT / "dist"


@asynccontextmanager
async def lifespan(app: FastAPI):
    db.initialize()
    collector.start()
    yield
    collector.stop()


app = FastAPI(title="Enrich Control Centre", version="2.0.0", lifespan=lifespan)

PLANT_MAPPING_FILE = settings.scada_mapping_workbook if settings.scada_mapping_workbook.exists() else ROOT / "Control_Centre_plantwise_data_mapping.xlsx"
PLANT_SITE_ALIASES = {
    "Bhokar - I": "Bhokar", "Polangal": "NLC Poolangal", "Rajgir": "BEL1MW",
    "Muradnagar": "BEL2MW", "Nagdha": "PGCIL",
}
UMRI_MCR = {"lat": 19.087861, "lon": 77.696167}
UMRI_LIVE_PLANTS = {
    "WIF": {"collection": "U2_WIF_LIVE", "lat": 19.088278, "lon": 77.696833},
    "WHF-1": {"collection": "U1_WHF_LIVE", "lat": 19.089361, "lon": 77.699028},
    "Haldiram": {"collection": "U5_Haldiram_LIVE", "lat": 19.093889, "lon": 77.713417},
    "Klassic Wheels": {"collection": "U3_Klassic_LIVE", "lat": 19.097806, "lon": 77.712306},
    "Marvelous": {"collection": "U4_Marvelous_LIVE", "lat": 19.098472, "lon": 77.710806},
    "Parakh": {"collection": "U6_Parakh_LIVE", "lat": 19.100056, "lon": 77.720306},
    "WHF-2": {"collection": "U9_WHF_2_LIVE", "lat": 19.104389, "lon": 77.720861},
    "PV Sons": {"collection": "U7_PV_Sons_LIVE", "lat": 19.105417, "lon": 77.720389},
}


def _coordinate(value, limit):
    if value is None:
        return None
    coordinate = float(value)
    while abs(coordinate) > limit:
        coordinate /= 10
    return coordinate


@app.get("/api/plant-mapping")
def plant_mapping():
    if not PLANT_MAPPING_FILE.exists():
        raise HTTPException(404, detail="Plant mapping workbook is unavailable")
    sheet = load_workbook(PLANT_MAPPING_FILE, data_only=True, read_only=True).active
    sites: dict[str, list[dict]] = {}
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row[4]:
            continue
        workbook_site = str(row[4]).strip()
        site_name = PLANT_SITE_ALIASES.get(workbook_site, workbook_site)
        plant_name = str(row[2]).strip() if row[2] else f"{site_name} Plant"
        has_live_columns = len(row) > 8
        collection = str(row[8] or "").strip() if has_live_columns else ""
        plants = sites.setdefault(site_name, [])
        plants.append({
            "id": f"mapping-{row_index}", "customerName": str(row[1] or "").strip(),
            "plantName": plant_name, "state": str(row[3] or "").strip(), "siteName": site_name,
            "ac": float(row[5] or 0), "dc": float(row[6] or 0),
            "commissioningDate": row[7].date().isoformat() if isinstance(row[7], datetime) else str(row[7] or ""),
            "communicationIssue": False,
            "collection": collection,
            "lat": _coordinate(row[9], 90) if len(row) > 9 else None,
            "lon": _coordinate(row[10], 180) if len(row) > 10 else None,
            **({"mcrLat": UMRI_MCR["lat"], "mcrLon": UMRI_MCR["lon"]} if site_name == "Umri" else {}),
        })
    # The live-collection workbook intentionally contains only SCADA-connected
    # plants. Merge non-live sites from the commissioned plant master so sites
    # such as Mandrup still support site -> plant drill-down.
    commissioned_workbook = ROOT / "Control_Centre_plantwise_data_mapping.xlsx"
    if commissioned_workbook.exists() and commissioned_workbook.resolve() != PLANT_MAPPING_FILE.resolve():
        live_mapped_sites = set(sites)
        commissioned_sheet = load_workbook(commissioned_workbook, data_only=True, read_only=True).active
        for row_index, row in enumerate(commissioned_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[4]:
                continue
            workbook_site = str(row[4]).strip()
            site_name = PLANT_SITE_ALIASES.get(workbook_site, workbook_site)
            if site_name in live_mapped_sites:
                continue
            sites.setdefault(site_name, []).append({
                "id": f"commissioned-{row_index}", "customerName": str(row[1] or "").strip(),
                "plantName": str(row[2] or f"{site_name} Plant").strip(),
                "state": str(row[3] or "").strip(), "siteName": site_name,
                "ac": float(row[5] or 0), "dc": float(row[6] or 0),
                "commissioningDate": row[7].date().isoformat() if isinstance(row[7], datetime) else str(row[7] or ""),
                "communicationIssue": False, "collection": "", "lat": None, "lon": None,
            })
    # The legacy workbook has the PGCIL capacity row but no collection/coordinate
    # columns. Keep it available to the common site -> plant -> inverter drill-down.
    if "PGCIL" not in sites or not any(plant.get("collection") for plant in sites["PGCIL"]):
        sites["PGCIL"] = [{
            "id": "mapping-pgcil-live", "customerName": "Power Grid Corporation of India Limited (PGCIL)",
            "plantName": "PGCIL", "state": "Madhya Pradesh", "siteName": "PGCIL",
            "ac": 85.0, "dc": 107.957, "commissioningDate": "2025-06-05",
            "communicationIssue": False, "collection": "PGCIL_LIVE",
            "lat": 23.41134, "lon": 75.4809,
        }]
    return {"source": PLANT_MAPPING_FILE.name, "sites": sites}

WEATHER_HOURLY = ",".join([
    "temperature_2m", "relative_humidity_2m", "apparent_temperature",
    "precipitation", "rain", "weather_code", "cloud_cover",
    "wind_speed_10m", "wind_direction_10m", "wind_gusts_10m",
    "shortwave_radiation", "direct_radiation", "diffuse_radiation",
    "global_tilted_irradiance",
])
WEATHER_DAILY = ",".join([
    "weather_code", "temperature_2m_max", "temperature_2m_min",
    "apparent_temperature_max", "apparent_temperature_min", "precipitation_sum",
    "rain_sum", "precipitation_probability_max", "sunrise", "sunset",
    "sunshine_duration", "wind_speed_10m_max", "wind_gusts_10m_max",
    "shortwave_radiation_sum",
])
_incident_cache = {"expires": datetime.min, "rows": []}
_weather_current_cache: dict[tuple[float, float], tuple[datetime, dict]] = {}


class OperationalFeedEntry(BaseModel):
    event_type: str = Field(pattern="^(alarm|event)$")
    plant: str
    message: str
    severity: str
    source: str
    timestamp: datetime


class OperationalFeedBatch(BaseModel):
    entries: list[OperationalFeedEntry]


def _valid_plant(plant: str) -> str:
    names = {label.upper(): label for _, label, _ in TARGET_STATIONS}
    aliases = {
        "KARAJAGI": "ENRICH KARASGI", "KARAJGI": "ENRICH KARASGI",
        "KARJAGI": "ENRICH KARASGI", "UMRI": "ENRICH ENERGY HIRADGAON",
        "TULJAPUR": "ENRICH TULJAPUR", "MANDRUP": "ENRICH MANDRUP",
        "KUMBHARI": "ENRICH ENERGY LTD SOLAR PARK",
        "BHOKAR PHASE-1": "ENRICH ENERGY BHOKAR",
        "BHOKAR PHASE-2": "ENRICH SOLAR SERVICES (Narwat)",
    }
    names.update(aliases)
    try:
        return names[plant.upper()]
    except KeyError as exc:
        raise HTTPException(400, detail="Unknown SLDC plant") from exc


@app.get("/api/sldc/live")
def live(response: Response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return db.latest()


@app.get("/api/scada/live")
def scada_live(response: Response):
    """Latest summed INV1..INV20 readings, grouped by configured site."""
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    try:
        return scada_reader.latest()
    except Exception as exc:
        raise HTTPException(503, detail="SCADA MongoDB unavailable") from exc


@app.get("/api/scada/sites/{site_name}")
def scada_site_details(site_name: str, response: Response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    details = scada_reader.site_details(site_name)
    if details is None:
        raise HTTPException(404, detail="SCADA site is not configured")
    return details


@app.get("/api/scada/sites/{site_name}/plants/{collection_name}")
def scada_plant_details(site_name: str, collection_name: str, response: Response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    details = scada_reader.plant_details(site_name, collection_name)
    if details is None:
        raise HTTPException(404, detail="SCADA plant collection is not configured")
    return details


@app.get("/api/sldc/samples")
def samples(plant: str, start: datetime, end: datetime):
    if end <= start:
        raise HTTPException(400, detail="end must be later than start")
    return db.samples(_valid_plant(plant), start, end)


@app.get("/api/sldc/availability")
def availability(plant: str, start: datetime, end: datetime,
                 group_by: str = Query("day", pattern="^(day|month|none)$")):
    if end <= start:
        raise HTTPException(400, detail="end must be later than start")
    return db.availability(_valid_plant(plant), start, end, group_by)


@app.get("/api/sldc/fleet-availability")
def fleet_availability(start: datetime, end: datetime):
    if end <= start:
        raise HTTPException(400, detail="end must be later than start")
    return db.fleet_availability(start, end)


@app.get("/api/sldc/incidents/active")
def active_sldc_incidents(start: datetime, end: datetime):
    if end <= start:
        raise HTTPException(400, detail="end must be later than start")
    return db.active_communication_incidents(start, end)


@app.get("/api/sldc/generation")
def generation(plant: str, start: datetime, end: datetime,
               group_by: str = Query("day", pattern="^(day|month|none)$")):
    if end <= start:
        raise HTTPException(400, detail="end must be later than start")
    canonical = _valid_plant(plant)
    return db.generation_report([canonical], start, end, group_by)


@app.get("/api/sldc/communication")
def communication(plant: str, start: datetime, end: datetime):
    if end <= start:
        raise HTTPException(400, detail="end must be later than start")
    canonical = _valid_plant(plant)
    return db.communication_report([canonical], start, end)


def _excel_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list]):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="0B5C94")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(45, max(11, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[column[0].column_letter].width = width
    return sheet


@app.get("/api/sldc/report.xlsx")
def excel_report(plant: str, start: datetime, end: datetime):
    if end <= start:
        raise HTTPException(400, detail="end must be later than start")
    canonical = _valid_plant(plant)
    samples = db.samples(canonical, start, end)
    availability = db.availability(canonical, start, end, "day")
    generation = db.generation_report([canonical], start, end, "day")
    communication = db.communication_report([canonical], start, end)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary_rows = [
        ("Plant", canonical),
        ("From", start.strftime("%Y-%m-%d %H:%M:%S")),
        ("To", end.strftime("%Y-%m-%d %H:%M:%S")),
        ("Estimated generation (MWh)", round(sum(row["EstimatedGenerationMWh"] for row in generation), 3)),
        ("15-minute logs", len(samples)),
        ("Unavailable slots", sum(row["UnavailableSamples"] for row in availability)),
        ("Negative MW rule", "Displayed in logs; treated as 0 only in generation totals"),
    ]
    for row in summary_rows:
        summary.append(row)
    summary["A1"].font = Font(bold=True, color="38BAFF")
    summary.column_dimensions["A"].width = 31
    summary.column_dimensions["B"].width = 62

    _excel_sheet(workbook, "Daily Performance",
        ["Date", "Estimated Generation (MWh)", "Average MW", "Minimum MW", "Maximum MW",
         "Available Samples", "Expected Samples", "Availability (%)"],
        [[row["Period"], row["EstimatedGenerationMWh"], row["AverageMW"], row["MinimumMW"],
          row["MaximumMW"], row["AvailableSamples"], row["ExpectedSamples"],
          row["AvailabilityPercent"]] for row in generation])
    _excel_sheet(workbook, "Communication Issues",
        ["Issue Start", "Issue End", "Duration (minutes)", "Lost 15-min Slots", "Issue"],
        [[row["StartTime"], row["EndTime"], row["DurationMinutes"], row["LostSamples"], row["Issue"]]
         for row in communication])
    _excel_sheet(workbook, "15-minute Logs",
        ["Sample Time", "Power (MW)", "Communication", "SLDC Status", "Issue Detail",
         "Source Timestamp", "Collected At"],
        [[row["SampleTime"], row["MW"], "Available" if row["IsAvailable"] else "Unavailable",
          row["Status"], row["CommunicationIssue"] or row["DashboardStatus"],
          row["SourceTimestamp"], row["CollectedAt"]] for row in samples])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    safe_plant = re.sub(r"[^A-Za-z0-9]+", "_", canonical).strip("_")
    filename = f"{safe_plant}_{start:%Y%m%d}_{end:%Y%m%d}_SLDC_Report.xlsx"
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"',
                 "Cache-Control": "no-store"})


def _weather_response(url: str, params: dict):
    try:
        result = requests.get(url, params=params, timeout=20)
        result.raise_for_status()
        return result.json()
    except (requests.RequestException, ValueError) as exc:
        detail = "Live weather provider unavailable"
        if getattr(exc, "response", None) is not None:
            try:
                detail = exc.response.json().get("reason", detail)
            except ValueError:
                pass
        raise HTTPException(502, detail=detail) from exc


@app.get("/api/weather/current")
def weather_current(lat: float = Query(ge=-90, le=90), lon: float = Query(ge=-180, le=180)):
    """Current conditions fallback used when the primary weather API is rate limited."""
    cache_key = (round(lat, 4), round(lon, 4))
    cached = _weather_current_cache.get(cache_key)
    if cached and cached[0] > datetime.now():
        return cached[1]
    try:
        result = requests.get(
            "https://api.met.no/weatherapi/locationforecast/2.0/compact",
            params={"lat": lat, "lon": lon}, timeout=20,
            headers={"User-Agent": "Enrich-Control-Centre/2.0 contact@enrichenergy.com"},
        )
        result.raise_for_status()
        point = result.json()["properties"]["timeseries"][0]
        details = point["data"]["instant"]["details"]
        next_hour = point["data"].get("next_1_hours", {})
        symbol = next_hour.get("summary", {}).get("symbol_code", "")
        weather_code = 95 if "thunder" in symbol else 61 if "rain" in symbol else 45 if "fog" in symbol else 3 if "cloudy" in symbol else 2 if "partlycloudy" in symbol else 0
        local_hour = (datetime.utcnow() + timedelta(hours=5, minutes=30)).hour + (datetime.utcnow().minute / 60)
        solar_curve = max(0.0, math.sin(((local_hour - 6) / 12) * math.pi))
        cloud = float(details.get("cloud_area_fraction", 0))
        cloud_factor = max(0.18, 1 - 0.72 * cloud / 100)
        estimated_gti = round(950 * solar_curve * cloud_factor)
        # Integral of a representative 12-hour solar curve. This supplies the
        # full-day irradiation value even when the current reading is at night.
        daily_gti_kwh_m2 = round(950 * cloud_factor * (24 / math.pi) / 1000, 2)
        payload = {"current": {
            "time": point["time"], "temperature_2m": details.get("air_temperature"),
            "relative_humidity_2m": details.get("relative_humidity"),
            "precipitation": next_hour.get("details", {}).get("precipitation_amount", 0),
            "rain": next_hour.get("details", {}).get("precipitation_amount", 0), "showers": 0,
            "weather_code": weather_code, "wind_speed_10m": round(float(details.get("wind_speed", 0)) * 3.6, 1),
            "shortwave_radiation": estimated_gti, "global_tilted_irradiance": estimated_gti,
            "daily_gti_kwh_m2": daily_gti_kwh_m2,
            "source_name": "MET Norway Locationforecast fallback",
        }}
        _weather_current_cache[cache_key] = (datetime.now() + timedelta(minutes=10), payload)
        return payload
    except (requests.RequestException, KeyError, IndexError, ValueError) as exc:
        raise HTTPException(502, detail="Fallback weather provider unavailable") from exc


@app.get("/api/weather/forecast")
def weather_forecast(lat: float = Query(ge=-90, le=90), lon: float = Query(ge=-180, le=180)):
    current = WEATHER_HOURLY + ",is_day,showers"
    return _weather_response("https://api.open-meteo.com/v1/forecast", {
        "latitude": lat, "longitude": lon, "timezone": "Asia/Kolkata",
        "current": current, "hourly": WEATHER_HOURLY, "daily": WEATHER_DAILY,
        "forecast_days": 7,
    })


@app.get("/api/weather/history")
def weather_history(lat: float = Query(ge=-90, le=90), lon: float = Query(ge=-180, le=180),
                    start_date: date = Query(), end_date: date = Query()):
    if end_date < start_date:
        raise HTTPException(400, detail="end_date must be on or after start_date")
    if (end_date - start_date).days > 366:
        raise HTTPException(400, detail="Historical range cannot exceed 366 days")
    return _weather_response("https://archive-api.open-meteo.com/v1/archive", {
        "latitude": lat, "longitude": lon, "timezone": "Asia/Kolkata",
        "start_date": start_date.isoformat(), "end_date": end_date.isoformat(),
        "hourly": WEATHER_HOURLY,
    })


@app.post("/api/operations/feed")
def save_operational_feed(batch: OperationalFeedBatch):
    now = datetime.now()
    rows = []
    for entry in batch.entries[:100]:
        event_time = entry.timestamp
        if entry.event_type == "alarm" and entry.source == "MH SLDC":
            try:
                canonical = _valid_plant(entry.plant)
                if now >= _incident_cache["expires"]:
                    start = (now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
                    _incident_cache["rows"] = db.active_communication_incidents(start, now)
                    _incident_cache["expires"] = now + timedelta(seconds=5)
                incident = next((row for row in _incident_cache["rows"] if row["Plant"] == canonical), None)
                if incident and entry.message.startswith(incident["Issue"]):
                    event_time = datetime.fromisoformat(incident["StartTime"])
            except HTTPException:
                pass
        raw = f"{entry.event_type}|{entry.plant}|{entry.message}|{entry.severity}|{entry.source}|{event_time.isoformat()}"
        rows.append({
            "EventKey": hashlib.sha256(raw.encode("utf-8")).hexdigest(),
            "EventType": entry.event_type,
            "PlantName": entry.plant,
            "Message": entry.message,
            "Severity": entry.severity,
            "SourceName": entry.source,
            "EventTime": event_time,
            "CreatedAt": now,
        })
    return {"stored": db.save_operational_events(rows), "received": len(rows)}


@app.get("/api/operations/logs")
def operational_logs(start: datetime, end: datetime,
                     event_type: str = Query("all", pattern="^(all|alarm|event)$"),
                     plant: str | None = None, severity: str | None = None,
                     limit: int = Query(500, ge=1, le=2000)):
    if end <= start:
        raise HTTPException(400, detail="end must be later than start")
    return db.operational_events(start, end, event_type, plant, severity, limit)


@app.get("/health")
def health():
    return {
        "status": "ok",
        "database": db.kind,
        "sampleDatabase": db.sample_kind,
        "sampleCollection": f"{settings.mongodb_database}.{settings.mongodb_collection}" if settings.mongodb_uri else "SLDC_DB",
        "collectorRunning": bool(collector.thread and collector.thread.is_alive()),
        "plantMappingAvailable": PLANT_MAPPING_FILE.exists(),
    }


if DIST.exists():
    app.mount("/assets", StaticFiles(directory=DIST / "assets"), name="assets")

    @app.get("/{path:path}", include_in_schema=False)
    def react_app(path: str):
        candidate = (DIST / path).resolve()
        if path and candidate.is_file() and DIST.resolve() in candidate.parents:
            return FileResponse(candidate)
        return FileResponse(DIST / "index.html", headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
        })
else:
    @app.get("/", include_in_schema=False)
    def missing_build():
        return {"message": "Run npm run build before starting the integrated server."}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("server:app", host=settings.api_host, port=settings.api_port, reload=False)
